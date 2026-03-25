#!/usr/bin/env python3
"""
Real-time sync from SQLite DB to master Excel spreadsheet.
Writes new leads and status updates back to the Excel file.
"""
import os
import sqlite3
import openpyxl
from datetime import datetime

DB_PATH = 'lead_pipeline.db'
SPREADSHEET_PATH = 'C:/Users/TrifectaAgent/.openclaw/media/inbound/1921d37d-3f7b-4702-8f7f-8919ac87b967.xlsx'

def log_action(message):
    """Log to console and file."""
    timestamp = datetime.now().isoformat()
    log_msg = f"[{timestamp}] {message}"
    print(log_msg)

def get_all_leads():
    """Get all leads from SQLite DB."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, name, email, phone, source, status, 
                   initial_question, program_interest, notes,
                   created_at, updated_at
            FROM leads
            WHERE status NOT IN ('DELETED', 'ARCHIVED')
            ORDER BY created_at DESC
        """)
        
        leads = [dict(row) for row in cursor.fetchall()]
        conn.close()
        log_action(f"Retrieved {len(leads)} leads from DB")
        return leads
        
    except Exception as e:
        log_action(f"ERROR fetching leads: {str(e)}")
        return []

def reverse_status_map(db_status):
    """Convert DB status back to spreadsheet status."""
    status_map = {
        'INQUIRY_RECEIVED': 'Awaiting Response',
        'PROGRAM_INFO_SENT': 'Program Info Sent',
        'CUSTOMER_REPLIED': 'Customer Replied',
        'BUSINESS_RESPONDED': 'Business Responded',
        'NEEDS_HUMAN_REVIEW': 'Awaiting Response',
        'CONSULTATION_COMPLETED': 'Consultation Completed',
        'FOLLOWUP_SENT': 'Follow-up Sent',
        'NO_RESPONSE': 'No Response',
        'AWAITING_RESPONSE': 'Awaiting Response',
        'AWAITING_SEND': 'Response Sent',
        'AWAITING_EMAIL': 'Program Info Sent',
    }
    return status_map.get(db_status, db_status)

def sync_to_spreadsheet():
    """Write all leads to spreadsheet."""
    try:
        leads = get_all_leads()
        if not leads:
            log_action("No leads to sync")
            return
        
        # Load spreadsheet
        wb = openpyxl.load_workbook(SPREADSHEET_PATH)
        ws = wb['Sheet1']
        
        # Column mapping (0-indexed)
        COL = {
            'date': 1,
            'name': 2,
            'email': 3,
            'phone': 4,
            'source': 5,
            'question': 6,
            'date_resp': 7,
            'status': 8,
            'followup': 9,
            'program': 10,
            'notes': 11,
        }
        
        # Clear existing data (keep headers)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 13):
                ws.cell(row=row_idx, column=col_idx).value = None
        
        # Write all leads
        for idx, lead in enumerate(leads, start=2):
            try:
                # Parse date
                created_date = lead.get('created_at', '')
                if created_date:
                    created_date = created_date[:10]  # YYYY-MM-DD
                
                # Map values
                ws.cell(row=idx, column=COL['date'] + 1).value = created_date
                ws.cell(row=idx, column=COL['name'] + 1).value = lead.get('name', '')
                ws.cell(row=idx, column=COL['email'] + 1).value = lead.get('email', '')
                ws.cell(row=idx, column=COL['phone'] + 1).value = lead.get('phone', '')
                ws.cell(row=idx, column=COL['source'] + 1).value = lead.get('source', 'GODADDY')
                ws.cell(row=idx, column=COL['question'] + 1).value = lead.get('initial_question', '')
                ws.cell(row=idx, column=COL['status'] + 1).value = reverse_status_map(lead.get('status', ''))
                ws.cell(row=idx, column=COL['program'] + 1).value = lead.get('program_interest', '')
                ws.cell(row=idx, column=COL['notes'] + 1).value = lead.get('notes', '')
                
            except Exception as e:
                log_action(f"ERROR writing row {idx}: {str(e)}")
        
        # Save spreadsheet
        wb.save(SPREADSHEET_PATH)
        log_action(f"Synced {len(leads)} leads to spreadsheet")
        
    except Exception as e:
        log_action(f"ERROR syncing to spreadsheet: {str(e)}")

if __name__ == "__main__":
    log_action("Starting DB to spreadsheet sync...")
    sync_to_spreadsheet()
    log_action("Sync complete.")
